import calendar
from datetime import datetime, timedelta
import json

import openpyxl
import stripe
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.template import loader
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy, reverse
from django.views.decorators.csrf import csrf_exempt

from Insportify import settings
from .forms import MultiStepForm, UserForm, AvailabilityForm
from .models import master_table, Individual, Organization, Venues, SportsCategory, SportsType, Order, User, \
    Availability
from django.views.generic import FormView


@login_required
def multistep(request):
    if request.method == "POST":
        form = MultiStepForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Event Created')
            return render(request, 'EventsApp/multi_step.html', {'form': form})
    else:
        form = MultiStepForm
        if 'submitted' in request.GET:
            submitted = True

    return render(request, 'EventsApp/multi_step.html', {'form': form})


@login_required
def all_events(request):
    event_list = master_table.objects.all()
    return render(request, 'EventsApp/event_list.html', {'event_list': event_list})


@login_required
def event_by_id(request, event_id):
    event = master_table.objects.get(pk=event_id)
    context = {
        'STRIPE_PUBLISHABLE_KEY': settings.STRIPE_PUBLISHABLE_KEY,
        'event': event
    }
    return render(request, 'EventsApp/event_detail.html', context)


@login_required
def user_profile(request):
    context = {
        'user': request.user
    }
    sports_category = SportsCategory.objects.all()
    sports_type = SportsType.objects.all()
    context['sports_category'] = sports_category
    context['sports_type'] = sports_type

    if request.method == "GET":
        individual = Individual.objects.get(user=request.user)
        # print(individual.__dict__)
        context['individual'] = individual
        return render(request, 'registration/individual_view.html', context)

    elif request.method == "POST":
        individual = Individual.objects.filter(user=request.user)
        response = request.POST.dict()
        print(response)
        if individual.exists():
            individual = Individual.objects.get(user=request.user)
            individual.user = request.user
            individual.first_name = response["first_name"].strip()
            individual.last_name = response["last_name"].strip()
            individual.phone = response["mobile"].strip()
            individual.email = response["contact_email"].strip()
            individual.provider = response["provider"].strip()
            individual.dob = response["dob"].strip()
            individual.concussion = response["is_concussion"].strip()
            individual.participation_interest = response["interest_gender"].strip()
            individual.city = response["city"].strip()
            individual.province = response["province"].strip()
            individual.country = response["Country"].strip()
            individual.sports_category = response["sport_category"].strip()
            individual.sports_type = response["sport_type"].strip()
            individual.sports_position = response["position"].strip()
            individual.sports_skill = response["skill"].strip()
            individual.save()
            context['individual'] = individual
        else:
            obj = Individual()
            obj.user = request.user
            obj.first_name = response["first_name"].strip()
            obj.last_name = response["last_name"].strip()
            obj.phone = response["mobile"].strip()
            obj.email = response["contact_email"].strip()
            obj.provider = response["provider"].strip()
            obj.dob = response["dob"].strip()
            obj.concussion = response["is_concussion"].strip()
            obj.participation_interest = response["interest_gender"].strip()
            obj.city = response["city"].strip()
            obj.province = response["province"].strip()
            obj.country = response["Country"].strip()
            obj.sports_category = response["sport_category"].strip()
            obj.sports_type = response["sport_type"].strip()
            obj.sports_position = response["position"].strip()
            obj.sports_skill = response["skill"].strip()
            obj.save()
            context['individual'] = obj
        messages.success(request, 'Individual details updated!')

    return render(request, 'registration/individual_view.html', context)


def get_selected_sports_type(request):
    data = {}
    if request.method == "POST":
        selected_category = request.POST['selected_category_text']
        try:
            print(selected_category)
            selected_type = SportsType.objects.filter(sports_category__sports_catgeory_text=selected_category)
            print(selected_type)
        except Exception:
            data['error_message'] = 'error'
            return JsonResponse(data)
        return JsonResponse(list(selected_type.values('pk', 'sports_type_text')), safe=False)


def get_sports_category(request):
    data = {}
    if request.method == "GET":
        try:
            selected_type = SportsCategory.objects.all()
        except Exception:
            data['error_message'] = 'error'
            return JsonResponse(data)
        return JsonResponse(list(selected_type.values('pk', 'sports_catgeory_text')), safe=False)


@login_required
def organization_profile(request):
    context = {
        'user': request.user
    }
    if request.method == "GET":
        organization = Organization.objects.get(user=request.user)
        print(organization.__dict__)
        context['organization'] = organization
        return render(request, 'registration/organization_view.html', context)

    if request.method == "POST":
        organization = Organization.objects.filter(user=request.user)
        response = request.POST.dict()
        if organization.exists():
            organization = Organization.objects.get(user=request.user)
            organization.user = request.user
            organization.type_of_organization = response["type_of_organization"].strip()
            organization.organization_name = response["company_name"].strip()
            organization.parent_organization_name = response["parent_organization"].strip()
            organization.registration_no = response["registration"].strip()
            organization.street = response["street_name"].strip()
            organization.city = response["city"].strip()
            organization.province = response["province"].strip()
            organization.country = response["country"].strip()
            organization.postal_code = response["postal_code"].strip()
            organization.email = response["email"].strip()
            organization.phone = response["phone"].strip()
            organization.website = response["website"].strip()
            organization.gender_focus = response["gender"].strip()
            organization.age_group = response["age_group"].strip()
            organization.save()
            context['organization'] = organization
        else:
            obj = Organization()
            obj.user = request.user
            obj.type_of_organization = response["type_of_organization"].strip()
            obj.organization_name = response["company_name"].strip()
            obj.parent_organization_name = response["parent_organization"].strip()
            obj.registration_no = response["registration"].strip()
            obj.street = response["street_name"].strip()
            obj.city = response["city"].strip()
            obj.province = response["province"].strip()
            obj.country = response["country"].strip()
            obj.postal_code = response["postal_code"].strip()
            obj.email = response["email"].strip()
            obj.phone = response["phone"].strip()
            obj.website = response["website"].strip()
            obj.gender_focus = response["gender"].strip()
            obj.age_group = response["age_group"].strip()
            obj.save()
            context['organization'] = obj
        messages.success(request, 'Organization details updated!')
    return render(request, 'registration/organization_view.html', context)


class UserProfileView(FormView):
    form_class = UserForm
    template_name = 'EventsApp/user_profile.html'
    success_url = reverse_lazy('multistep')


def home(request):
    sports = SportsCategory.objects.values('pk', 'sports_catgeory_text')
    venues = Venues.objects.values('pk', 'vm_name')
    load_venues_excel()
    events = master_table.objects.all()
    recommended_events = get_recommended_events(request)

    if request.GET.get('events_types'):
        selected_events_types = request.GET.get('events_types')
        # events = events.filter(Q(event_type__icontains=selected_events_types))
        events = events.filter(event_type=selected_events_types)

    if request.GET.get('sports'):
        selected_sports = request.GET.get('sports')
        events = events.filter(sport_type=selected_sports)

    if request.GET.get('venues'):
        selected_venues = request.GET.get('venues')
        events = events.filter(venue=selected_venues)

    if request.GET.get('date_range'):
        selected_date = request.GET.get('date_range')
        selected_date = datetime.strptime(selected_date.strip(), '%Y-%m-%d').date()
        events = get_events_by_date(events, selected_date)

    recommended_events = [recommended_events[i:i + 3] for i in range(0, len(recommended_events), 3)]
    events = [events[i:i + 3] for i in range(0, len(events), 3)]
    context = {
        'sports_list': sports,
        'venues_list': venues,
        'events': events,
        'recommended_events': recommended_events
    }
    html_template = loader.get_template('EventsApp/home.html')
    return HttpResponse(html_template.render(context, request))


def get_recommended_events(request):
    user = User.objects.get(username=request.user.username)
    user_avaiability = Availability.objects.filter(user=user)
    events = master_table.objects.all()
    week_days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    recommended_events=set()
    for event in events:
        time = event.datetimes.split("-")
        start_datetime = datetime.strptime(time[0].strip(), '%m/%d/%Y %I:%M %p')
        end_datetime = datetime.strptime(time[-1].strip(), '%m/%d/%Y %I:%M %p')

        for i in range((end_datetime - start_datetime).days):
            # print(i, calendar.day_name[(start_time + timedelta(days=i+1)).weekday()])
            days_between = calendar.day_name[(start_datetime + timedelta(days=i + 1)).weekday()]
            for avail in user_avaiability:
                if week_days[avail.day_of_week - 1] == days_between:
                    # print("day match",week_days[avail.day_of_week-1], days_between)
                    if avail.start_time <= end_datetime.time() or avail.end_time >= start_datetime.time():
                        recommended_events.add(event)

    # print(list(recommended_events))

    return list(recommended_events)


def get_events_by_date(events, selected_date):
    for event in events:
        time = event.datetimes.split("-")
        start_date = datetime.strptime(time[0].strip(), '%m/%d/%Y %I:%M %p').date()
        end_date = datetime.strptime(time[-1].strip(), '%m/%d/%Y %I:%M %p').date()
        if selected_date < start_date or selected_date > end_date:
            events = events.exclude(pk=event.id)

    return events


@csrf_exempt
def create_checkout_session(request, id):
    event = get_object_or_404(master_table, pk=id)
    unit_amount = round(event.position_cost)
    stripe.api_key = settings.STRIPE_SECRET_KEY
    try:
        checkout_session = stripe.checkout.Session.create(
            payment_method_types=['card'],
            line_items=[
                {
                    'price_data': {
                        'currency': 'usd',
                        'product_data': {
                            'name': event.event_title,
                        },
                        'unit_amount': int(unit_amount),
                    },
                    'quantity': 1,
                }
            ],
            mode='payment',
            success_url=request.build_absolute_uri(
                reverse('EventsApp:payment-success')) + "?session_id={CHECKOUT_SESSION_ID}",
            cancel_url=request.build_absolute_uri(reverse('EventsApp:payment-cancel')),
        )

        order = Order()
        order.customer = User.objects.get(username=request.user.username)
        order.event = event
        order.order_date = datetime.datetime.now()
        order.order_amount = int(unit_amount)
        order.save()

        return JsonResponse({'sessionId': checkout_session.id})

    except Exception as e:
        return JsonResponse({'error': str(e)})


def paymentSuccess(request):
    print("Success ho gaya")
    context = {
        "payment_status": "success"
    }
    return render(request, "EventsApp/confirmation.html", context)


def paymentCancel(request):
    print("cancel ho gaya")
    context = {
        "payment_status": "fail"
    }
    return render(request, "EventsApp/confirmation.html", context)


@login_required
def add_availability(request):
    context = {}
    form = AvailabilityForm(request.POST or None,
                            instance=Availability(),
                            initial={'user': request.user})

    context['form'] = form
    user = User.objects.get(username=request.user.username)
    user_avaiability = Availability.objects.filter(user=user)
    get_day_of_week(user_avaiability)

    context["user_availability"] = user_avaiability

    if request.POST:
        if form.is_valid():
            obj = Availability(user=user,
                               day_of_week=form.cleaned_data['day_of_week'],
                               start_time=form.cleaned_data['start_time'],
                               end_time=form.cleaned_data['end_time'])
            obj.save()
            user_avaiability = Availability.objects.filter(user=user)
            get_day_of_week(user_avaiability)
            context["user_availability"] = user_avaiability
            messages.success(request, "New Availability Added!")
        else:
            print(form.errors)

    return render(request, "EventsApp/add_availability.html", context)


def get_day_of_week(user_avaiability):
    for avail in user_avaiability:
        if avail.day_of_week == 1:
            avail.day_of_week = "Monday"
        if avail.day_of_week == 2:
            avail.day_of_week = "Tuesday"
        if avail.day_of_week == 3:
            avail.day_of_week = "Wednesday"
        if avail.day_of_week == 4:
            avail.day_of_week = "Thursday"
        if avail.day_of_week == 5:
            avail.day_of_week = "Friday"
        if avail.day_of_week == 6:
            avail.day_of_week = "Saturday"
        if avail.day_of_week == 7:
            avail.day_of_week = "Sunday"


def load_venues_excel():
    path = "./venue.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    # Venues.objects.all().delete()
    for i in range(1, sheet_obj.max_row + 1):
        if sheet_obj.cell(row=i, column=5).value.strip() == "ON":
            vm_name = sheet_obj.cell(row=i, column=1).value.strip()
            vm_venue_description = sheet_obj.cell(row=i, column=2).value.strip()
            vm_venue_street = sheet_obj.cell(row=i, column=3).value.strip()
            vm_venuecity = sheet_obj.cell(row=i, column=4).value.strip()
            vm_venue_province = sheet_obj.cell(row=i, column=5).value.strip()
            vm_venue_country = sheet_obj.cell(row=i, column=6).value.strip()
            vm_venue_zip = sheet_obj.cell(row=i, column=7).value.strip()
            obj = Venues(vm_name=vm_name, vm_venue_description=vm_venue_description, vm_venue_street=vm_venue_street,
                         vm_venuecity=vm_venuecity, vm_venue_province=vm_venue_province,
                         vm_venue_country=vm_venue_country, vm_venue_zip=vm_venue_zip)
            # obj.save()


@login_required
def delete_by_id(request, event_id):
    try:
        event = master_table.objects.get(pk=event_id)
        event.delete()
        messages.success(request, "Event removed successfully!")

    except:
        print("Some error occurred!")

    return redirect('EventsApp:list-events')


@login_required
def delete_availability(request, id):
    try:
        avail = Availability.objects.get(pk=id)
        avail.delete()
        messages.success(request, "Availability removed successfully!")

    except:
        print("Some error occurred!")

    return redirect('EventsApp:add_availability')
